from openpyxl import Workbook

from app.ingestion.inspectors.header_detector import HeaderDetector
from app.ingestion.models.sheet_inspection import SheetInspection


class SheetDetector:

    def detect(
        self,
        workbook: Workbook
    ) -> SheetInspection:

        best_inspection = None

        for sheet in workbook.worksheets:

            inspection = self._score_sheet(sheet)

            if (
                best_inspection is None
                or inspection.total_score > best_inspection.total_score
            ):
                best_inspection = inspection

        return best_inspection

    def _score_sheet(
        self,
        sheet
    ) -> SheetInspection:

        row_score = min(sheet.max_row / 1000, 1.0) * 40

        column_score = min(sheet.max_column / 20, 1.0) * 15

        rows = [
            list(row)
            for row in sheet.iter_rows(
                min_row=1,
                max_row=20,
                values_only=True
            )
        ]

        header = HeaderDetector().detect(rows)

        header_score = header.confidence * 20

        penalty_score = 0

        if sheet.max_row < 5:
            penalty_score = -30

        total = (
            row_score
            + column_score
            + header_score
            + penalty_score
        )

        return SheetInspection(
            sheet_name=sheet.title,
            total_score=round(total, 2),
            row_score=round(row_score, 2),
            column_score=round(column_score, 2),
            header_score=round(header_score, 2),
            penalty_score=penalty_score
        )