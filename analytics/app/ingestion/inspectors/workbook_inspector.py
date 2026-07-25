from io import BytesIO

from openpyxl import load_workbook

from app.ingestion.inspectors.sheet_detector import SheetDetector
from app.ingestion.inspectors.header_detector import HeaderDetector
from app.ingestion.inspectors.footer_detector import FooterDetector

from app.ingestion.models.workbook_inspection import WorkbookInspection


class WorkbookInspector:

    def inspect(
        self,
        file_bytes: bytes
    ) -> WorkbookInspection:

        workbook = load_workbook(
            BytesIO(file_bytes),
            read_only=True,
            data_only=True
        )

        sheet = SheetDetector().detect(workbook)

        worksheet = workbook[sheet.sheet_name]

        rows = []

        for row in worksheet.iter_rows(
            values_only=True
        ):
            rows.append(list(row))

        header = HeaderDetector().detect(rows)

        footer = FooterDetector().detect(rows)

        workbook.close()

        return WorkbookInspection(

            sheet=sheet,

            header=header,

            footer=footer,

            rows=rows

        )